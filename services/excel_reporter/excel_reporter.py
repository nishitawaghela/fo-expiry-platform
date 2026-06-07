"""
excel_reporter.py
Generates a formatted Thursday pre-market Excel report.
Reads from dbt mart views in Supabase PostgreSQL.

Sheets:
  1. Summary       — PCR, max pain, spot price, market sentiment
  2. OI by Strike  — bar chart of CE vs PE open interest per strike
  3. PCR Trend     — line chart of PCR across batches this week
  4. Max Pain      — full strike table with max pain highlighted
  5. Anomalies     — flagged strikes from Isolation Forest
"""

import os
import psycopg2
import psycopg2.extras
from datetime import datetime, timezone
from dotenv import load_dotenv

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL")

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE   = "1B2A4A"
ACCENT_GOLD = "D4A017"
LIGHT_GRAY  = "F2F2F2"
RED_FILL    = "FF4444"
GREEN_FILL  = "00AA44"
AMBER_FILL  = "FFA500"
WHITE       = "FFFFFF"


# ── DB helpers ────────────────────────────────────────────────────────────────
def fetch(query: str, params=None) -> list:
    conn = psycopg2.connect(DATABASE_URL)
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [dict(r) for r in rows]


def fetch_summary() -> dict:
    rows = fetch("SELECT * FROM mart_expiry_summary LIMIT 1")
    return rows[0] if rows else {}


def fetch_oi_by_strike() -> list:
    return fetch("SELECT * FROM mart_max_pain ORDER BY strike")


def fetch_pcr_trend() -> list:
    return fetch("""
        SELECT batch_id, pcr, spot_price, market_sentiment, computed_at
        FROM int_pcr_trend
        ORDER BY computed_at DESC
        LIMIT 20
    """)


def fetch_anomalies() -> list:
    """Read latest anomalies from oi_snapshots where oi_shift is extreme."""
    return fetch("""
        SELECT strike, option_type, open_interest, oi_shift,
               implied_volatility, batch_id
        FROM stg_oi_snapshots
        WHERE oi_shift != 0
        ORDER BY ABS(oi_shift) DESC
        LIMIT 20
    """)


# ── Style helpers ─────────────────────────────────────────────────────────────
def header_fill(color=DARK_BLUE):
    return PatternFill("solid", fgColor=color)


def cell_fill(color):
    return PatternFill("solid", fgColor=color)


def bold_white():
    return Font(bold=True, color=WHITE, size=11)


def bold_dark():
    return Font(bold=True, color=DARK_BLUE, size=11)


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill    = header_fill()
        cell.font    = bold_white()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border  = thin_border()


def style_data_row(ws, row: int, cols: int, alternate: bool = False):
    fill_color = LIGHT_GRAY if alternate else WHITE
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = cell_fill(fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
def build_summary_sheet(wb: Workbook, summary: dict):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value     = "F&O Expiry Analytics Report"
    title_cell.font      = Font(bold=True, size=16, color=WHITE)
    title_cell.fill      = header_fill(DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:B2")
    sub_cell = ws["A2"]
    sub_cell.value     = f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}"
    sub_cell.font      = Font(italic=True, size=10, color=DARK_BLUE)
    sub_cell.alignment = Alignment(horizontal="center")
    sub_cell.fill      = cell_fill(LIGHT_GRAY)

    # Metrics
    metrics = [
        ("Symbol",           summary.get("symbol", "NIFTY")),
        ("Expiry",           summary.get("expiry", "-")),
        ("Spot Price",       f"₹{summary.get('spot_price', 0):,.2f}"),
        ("PCR",              f"{summary.get('pcr', 0):.4f}"),
        ("Market Sentiment", summary.get("market_sentiment", "-").upper()),
        ("Max Pain Strike",  f"₹{summary.get('max_pain', 0):,.0f}"),
        ("Total CE OI",      f"{summary.get('total_ce_oi', 0):,}"),
        ("Total PE OI",      f"{summary.get('total_pe_oi', 0):,}"),
        ("Total Strikes",    summary.get("total_strikes", 0)),
        ("Last Updated",     str(summary.get("computed_at", "-"))[:19]),
    ]

    for i, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=i, column=1).value = label
        ws.cell(row=i, column=1).font  = bold_dark()
        ws.cell(row=i, column=1).fill  = cell_fill(LIGHT_GRAY)
        ws.cell(row=i, column=1).border = thin_border()

        val_cell = ws.cell(row=i, column=2)
        val_cell.value  = value
        val_cell.border = thin_border()
        val_cell.alignment = Alignment(horizontal="center")

        # Colour sentiment
        if label == "Market Sentiment":
            sentiment = summary.get("market_sentiment", "neutral")
            if sentiment == "bullish":
                val_cell.fill = cell_fill(GREEN_FILL)
                val_cell.font = Font(bold=True, color=WHITE)
            elif sentiment == "bearish":
                val_cell.fill = cell_fill(RED_FILL)
                val_cell.font = Font(bold=True, color=WHITE)
            else:
                val_cell.fill = cell_fill(AMBER_FILL)
                val_cell.font = Font(bold=True, color=WHITE)

        ws.row_dimensions[i].height = 22


# ── Sheet 2: OI by Strike ─────────────────────────────────────────────────────
def build_oi_chart_sheet(wb: Workbook, oi_data: list):
    ws = wb.create_sheet("OI by Strike")
    ws.sheet_view.showGridLines = False

    headers = ["Strike", "CE OI", "PE OI", "CE OI Shift", "PE OI Shift", "Max Pain"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    style_header_row(ws, 1, len(headers))

    for row_idx, record in enumerate(oi_data, start=2):
        values = [
            record["strike"],
            record["ce_oi"],
            record["pe_oi"],
            record["ce_oi_shift"],
            record["pe_oi_shift"],
            "← MAX PAIN" if record.get("is_max_pain_strike") else "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col).value = val
        style_data_row(ws, row_idx, len(headers), alternate=(row_idx % 2 == 0))

        # Highlight max pain row
        if record.get("is_max_pain_strike"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = cell_fill(ACCENT_GOLD)
                ws.cell(row=row_idx, column=col).font = Font(bold=True, color=DARK_BLUE)

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Add bar chart
    if len(oi_data) > 0:
        chart = BarChart()
        chart.type    = "col"
        chart.title   = "Open Interest by Strike"
        chart.y_axis.title = "Open Interest"
        chart.x_axis.title = "Strike Price"
        chart.style  = 10
        chart.width  = 30
        chart.height = 15

        last_row = len(oi_data) + 1
        ce_data  = Reference(ws, min_col=2, min_row=1, max_row=last_row)
        pe_data  = Reference(ws, min_col=3, min_row=1, max_row=last_row)
        strikes  = Reference(ws, min_col=1, min_row=2, max_row=last_row)

        chart.add_data(ce_data, titles_from_data=True)
        chart.add_data(pe_data, titles_from_data=True)
        chart.set_categories(strikes)

        chart.series[0].graphicalProperties.solidFill = "4472C4"  # CE blue
        chart.series[1].graphicalProperties.solidFill = "ED7D31"  # PE orange

        ws.add_chart(chart, f"H2")


# ── Sheet 3: PCR Trend ────────────────────────────────────────────────────────
def build_pcr_trend_sheet(wb: Workbook, pcr_data: list):
    ws = wb.create_sheet("PCR Trend")
    ws.sheet_view.showGridLines = False

    # Reverse to show chronological order
    pcr_data = list(reversed(pcr_data))

    headers = ["Batch", "PCR", "Spot Price", "Sentiment", "Computed At"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    style_header_row(ws, 1, len(headers))

    for row_idx, record in enumerate(pcr_data, start=2):
        values = [
            record["batch_id"][:20],
            round(record["pcr"], 4),
            record["spot_price"],
            record["market_sentiment"].upper(),
            str(record["computed_at"])[:19],
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col).value = val
        style_data_row(ws, row_idx, len(headers), alternate=(row_idx % 2 == 0))

        # Colour sentiment
        sentiment_cell = ws.cell(row=row_idx, column=4)
        if record["market_sentiment"] == "bullish":
            sentiment_cell.fill = cell_fill(GREEN_FILL)
            sentiment_cell.font = Font(bold=True, color=WHITE)
        elif record["market_sentiment"] == "bearish":
            sentiment_cell.fill = cell_fill(RED_FILL)
            sentiment_cell.font = Font(bold=True, color=WHITE)
        else:
            sentiment_cell.fill = cell_fill(AMBER_FILL)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # PCR line chart
    if len(pcr_data) > 1:
        chart = LineChart()
        chart.title   = "PCR Trend"
        chart.y_axis.title = "Put-Call Ratio"
        chart.x_axis.title = "Batch"
        chart.style  = 10
        chart.width  = 25
        chart.height = 12

        last_row = len(pcr_data) + 1
        pcr_vals = Reference(ws, min_col=2, min_row=1, max_row=last_row)
        chart.add_data(pcr_vals, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = "D4A017"
        chart.series[0].graphicalProperties.line.width = 20000

        ws.add_chart(chart, "G2")


# ── Sheet 4: Max Pain Table ───────────────────────────────────────────────────
def build_max_pain_sheet(wb: Workbook, oi_data: list):
    ws = wb.create_sheet("Max Pain")
    ws.sheet_view.showGridLines = False

    headers = ["Strike", "CE OI", "PE OI", "Total OI", "CE IV", "PE IV", "Is Max Pain"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    style_header_row(ws, 1, len(headers))

    for row_idx, record in enumerate(oi_data, start=2):
        values = [
            record["strike"],
            record["ce_oi"],
            record["pe_oi"],
            record["ce_oi"] + record["pe_oi"],
            round(record["ce_iv"], 2),
            round(record["pe_iv"], 2),
            "✓ MAX PAIN" if record.get("is_max_pain_strike") else "",
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col).value = val
        style_data_row(ws, row_idx, len(headers), alternate=(row_idx % 2 == 0))

        if record.get("is_max_pain_strike"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = cell_fill(ACCENT_GOLD)
                ws.cell(row=row_idx, column=col).font = Font(bold=True, color=DARK_BLUE)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


# ── Sheet 5: Anomalies ────────────────────────────────────────────────────────
def build_anomalies_sheet(wb: Workbook, anomalies: list):
    ws = wb.create_sheet("Anomalies")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value     = "⚠ Unusual OI Activity — Potential Institutional Positioning"
    title.font      = Font(bold=True, size=12, color=WHITE)
    title.fill      = cell_fill(RED_FILL)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Strike", "Option Type", "Open Interest", "OI Shift", "Implied Volatility", "Batch ID"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    style_header_row(ws, 2, len(headers))

    for row_idx, record in enumerate(anomalies, start=3):
        values = [
            record["strike"],
            record["option_type"],
            record["open_interest"],
            record["oi_shift"],
            round(record["implied_volatility"], 2),
            record["batch_id"][:20],
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col).value = val
        style_data_row(ws, row_idx, len(headers), alternate=(row_idx % 2 == 0))

        # Highlight large OI shifts in red
        oi_shift = record.get("oi_shift", 0)
        if abs(oi_shift) > 50000:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = cell_fill("FFE0E0")
                ws.cell(row=row_idx, column=col).font = Font(bold=True, color=RED_FILL)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


# ── Main ──────────────────────────────────────────────────────────────────────
def generate_report() -> str:
    print("Fetching data from PostgreSQL dbt views...")
    summary   = fetch_summary()
    oi_data   = fetch_oi_by_strike()
    pcr_data  = fetch_pcr_trend()
    anomalies = fetch_anomalies()

    print(f"Summary: {summary.get('expiry', 'N/A')} | PCR: {summary.get('pcr', 'N/A')} | Max Pain: {summary.get('max_pain', 'N/A')}")
    print(f"OI strikes: {len(oi_data)} | PCR trend batches: {len(pcr_data)} | Anomalies: {len(anomalies)}")

    wb = Workbook()

    build_summary_sheet(wb, summary)
    build_oi_chart_sheet(wb, oi_data)
    build_pcr_trend_sheet(wb, pcr_data)
    build_max_pain_sheet(wb, oi_data)
    build_anomalies_sheet(wb, anomalies)

    # Save report
    timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    output_path = os.path.abspath(f"fo_expiry_report_{timestamp}.xlsx")
    wb.save(output_path)

    print(f"\nReport saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    generate_report()